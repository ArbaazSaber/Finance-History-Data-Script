import os
import logging
import pandas as pd
import MetaTrader5 as mt5

from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side

# === CONFIG ===
FILE_PATH = "finance_data.xlsx"
TIMEFRAME = mt5.TIMEFRAME_D1
NUM_CANDLES = 100
TICKERS_SHEET_NAME = "Tickers"
DEFAULT_TICKER = "XAUUSDm"

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("market_data.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def initialize_mt5():
    """Initialize connection to MetaTrader 5."""
    try:
        if not mt5.initialize():
            error_msg = f"Failed to initialize MetaTrader 5. Error code: {mt5.last_error()}"
            logger.error(error_msg)
            raise RuntimeError(f"⚠️ {error_msg}")
            
        account_info = mt5.account_info()
        if account_info is None:
            mt5.shutdown()
            logger.error("No MT5 account info found.")
            raise RuntimeError("⚠️ No MT5 account info found.")
            
        logger.info(f"Connected to MT5 account: {account_info.login} ({account_info.server})")
        return True
    except Exception as e:
        logger.exception(f"Error initializing MT5: {e}")
        return False

def fetch_market_data(ticker: str) -> pd.DataFrame:
    """Fetch market data for the given ticker and return a DataFrame."""
    try:
        initialize_mt5()
        
        # Check if symbol exists
        symbol_info = mt5.symbol_info(ticker)
        if symbol_info is None:
            logger.warning(f"Symbol {ticker} not found in MT5")
            mt5.shutdown()
            return pd.DataFrame()
            
        # Fetch the data
        rates = mt5.copy_rates_from(ticker, TIMEFRAME, datetime.now(), NUM_CANDLES)
        mt5.shutdown()
        
        if rates is None or len(rates) == 0:
            logger.warning(f"⚠️ No data returned for {ticker}.")
            return pd.DataFrame()
            
        # Convert to DataFrame and process
        df = pd.DataFrame(rates)
        df["time"] = pd.to_datetime(df["time"], unit="s")
        
        # Add calculated fields
        df["range"] = df["high"] - df["low"]
        df["daily_change"] = df["close"] - df["open"]
        df["daily_change_pct"] = (df["daily_change"] / df["open"]) * 100
        
        logger.info(f"Successfully fetched {len(df)} candles for {ticker}")
        return df[["time", "open", "high", "low", "close", "range", 
                  "daily_change", "daily_change_pct", "tick_volume", "spread", "real_volume"]]
    except Exception as e:
        logger.exception(f"Error fetching data for {ticker}: {e}")
        try:
            mt5.shutdown()
        except:
            pass
        return pd.DataFrame()

def append_data_to_sheet(df: pd.DataFrame, ticker: str, file_path: str):
    """Append the given DataFrame to the Excel sheet named after the ticker."""
    if df.empty:
        logger.warning(f"No data to append for {ticker}")
        return
        
    try:
        sheet_name = f"{ticker} Data"
        
        if os.path.exists(file_path):
            book = load_workbook(file_path)
            if sheet_name in book.sheetnames:
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name, parse_dates=["time"])
                df = pd.concat([existing_df, df]).drop_duplicates(subset=["time"])
        
        df.sort_values("time", inplace=True)
        
        with pd.ExcelWriter(file_path, engine="openpyxl", 
                           mode="a" if os.path.exists(file_path) else "w", 
                           if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        logger.info(f"✅ {ticker} data written to '{sheet_name}' in {file_path}.")
    except Exception as e:
        logger.error(f"Error saving {ticker} data: {e}")

def ensure_workbook_exists(file_path: str, default_ticker: str):
    """Ensure the Excel file and Tickers sheet exist, with at least one default ticker."""
    try:
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = TICKERS_SHEET_NAME
            
            # Format header
            ws.append(["Ticker", "Description", "Last Updated"])
            for col in range(1, 4):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Add default ticker
            ws.append([default_ticker, "Gold vs USD", datetime.now().strftime("%Y-%m-%d %H:%M")])
            
            # Add summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet.append(["Market Data Summary"])
            summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            summary_sheet.append(["Last update:", datetime.now().strftime("%Y-%m-%d %H:%M")])
            
            wb.save(file_path)
            logger.info(f"📘 Created new workbook with default ticker: {default_ticker}")
        else:
            logger.debug(f"Workbook already exists: {file_path}")
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")

def read_tickers(file_path: str) -> list:
    """Return a list of tickers from the Tickers sheet."""
    try:
        tickers_df = pd.read_excel(file_path, sheet_name=TICKERS_SHEET_NAME)
        # Handle the case with or without Status column
        if "Status" in tickers_df.columns:
            active_tickers = tickers_df[tickers_df["Status"] == "Active"]["Ticker"].dropna().tolist()
            logger.info(f"Found {len(active_tickers)} active tickers")
            return active_tickers
        else:
            all_tickers = tickers_df["Ticker"].dropna().tolist()
            logger.info(f"Found {len(all_tickers)} tickers")
            return all_tickers
    except Exception as e:
        logger.error(f"❌ Error reading tickers: {e}")
        return []

def update_ticker_status(ticker: str, file_path: str, description: str = ""):
    """Update last updated timestamp for a ticker."""
    try:
        wb = load_workbook(file_path)
        ws = wb[TICKERS_SHEET_NAME]
        
        # Find ticker row
        ticker_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == ticker:
                ticker_row = row
                break
        
        if ticker_row:
            # Update existing ticker timestamp
            ws.cell(row=ticker_row, column=3).value = datetime.now().strftime("%Y-%m-%d %H:%M")
            if description and not ws.cell(row=ticker_row, column=2).value:
                ws.cell(row=ticker_row, column=2).value = description
        else:
            # Add new ticker
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1).value = ticker
            ws.cell(row=next_row, column=2).value = description
            ws.cell(row=next_row, column=3).value = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        wb.save(file_path)
        logger.info(f"Updated status for ticker: {ticker}")
    except Exception as e:
        logger.error(f"Error updating ticker status: {e}")

def update_overview_sheet(file_path: str, tickers: list, days: int = 15):
    """Creates an 'Overview' sheet showing the last `days` of data and charts for each ticker."""
    try:
        book = load_workbook(file_path)
        
        if "Overview" in book.sheetnames:
            del book["Overview"]
        
        overview_ws = book.create_sheet("Overview")
        
        # Add title
        overview_ws.cell(row=1, column=1, value=f"Market Data Overview (Last {days} Days)")
        overview_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        overview_ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        current_row = 4
        
        # Define styles
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Track price changes for summary
        price_changes = {}
        
        for ticker in tickers:
            sheet_name = f"{ticker} Data"
            if sheet_name not in book.sheetnames:
                logger.warning(f"⚠️ Sheet '{sheet_name}' not found. Skipping.")
                continue
            
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, parse_dates=["time"])
                if df.empty:
                    continue
                
                # Get the last N days of data
                df = df.sort_values("time").tail(days)
                
                # Calculate price changes if possible
                if len(df) > 1:
                    first_close = df.iloc[0]["close"]
                    last_close = df.iloc[-1]["close"]
                    price_change = last_close - first_close
                    price_change_pct = (price_change / first_close) * 100
                    price_changes[ticker] = {
                        "first_close": first_close,
                        "last_close": last_close,
                        "change": price_change,
                        "change_pct": price_change_pct
                    }
                
                # --- Write header for each ticker
                overview_ws.cell(row=current_row, column=1, value=f"Ticker: {ticker}")
                overview_ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                # Add data table - support both original and enhanced columns
                if "daily_change_pct" in df.columns:
                    columns_to_show = ["time", "open", "high", "low", "close", "daily_change_pct"]
                else:
                    columns_to_show = ["time", "open", "high", "low", "close"]
                
                for r in dataframe_to_rows(df[columns_to_show], index=False, header=True):
                    for col_idx, value in enumerate(r, start=1):
                        cell = overview_ws.cell(row=current_row, column=col_idx, value=value)
                        # Add basic formatting
                        if current_row == current_row - len(df) + 1:  # Header row
                            cell.font = Font(bold=True)
                            cell.fill = header_fill
                    current_row += 1
                
                # Format the Time column to 'DD-MMM'
                for row in overview_ws.iter_rows(min_row=current_row - len(df), 
                                                max_row=current_row - 1, 
                                                min_col=1, max_col=1):
                    for cell in row:
                        cell.number_format = 'DD-MMM'
                
                # Format percentage column if it exists
                if "daily_change_pct" in df.columns:
                    for row in overview_ws.iter_rows(min_row=current_row - len(df), 
                                                    max_row=current_row - 1, 
                                                    min_col=6, max_col=6):
                        for cell in row:
                            cell.number_format = '0.00%'
                            if cell.value > 0:
                                cell.font = Font(color="008800")
                            elif cell.value < 0:
                                cell.font = Font(color="CC0000")
                
                chart = LineChart()
                chart.title = f"{ticker} - Close Price (Last {days} Days)"
                chart.style = 13
                chart.width = 15
                chart.height = 7
                
                chart.legend = None
                chart.x_axis.title = "Date"
                chart.y_axis.title = "Close Price"
                
                chart.x_axis.number_format = 'DD-MMM'
                chart.x_axis.majorTimeUnit = "days"
                
                # Ensure axis lines and labels are shown
                chart.x_axis.majorTickMark = "out"
                chart.y_axis.majorTickMark = "out"
                chart.x_axis.tickLblPos = "nextTo"
                chart.y_axis.tickLblPos = "nextTo"
                chart.x_axis.visible = True
                chart.y_axis.visible = True
                
                chart.x_axis.majorGridlines = None
                chart.y_axis.majorGridlines = None
                
                # Reference for data
                header_row = current_row - len(df) - 1
                data_start = header_row + 1
                data_end = data_start + len(df) - 1
                
                values = Reference(overview_ws, min_col=5, min_row=header_row, max_row=data_end)  # Close column
                categories = Reference(overview_ws, min_col=1, min_row=data_start, max_row=data_end)  # Time column
                
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(categories)
                overview_ws.add_chart(chart, f"G{data_start}")
                
                current_row += 2  # spacing between tickers
            except Exception as e:
                logger.error(f"Error processing {ticker} for overview: {e}")
                continue
        
        # Create Summary table if we have price change data
        if price_changes and "Summary" in book.sheetnames:
            summary_ws = book["Summary"]
            summary_ws.cell(row=1, column=1, value="Market Data Summary")
            summary_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            summary_ws.cell(row=2, column=1, value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            
            # Price change summary table
            summary_ws.cell(row=4, column=1, value=f"{days}-Day Price Changes")
            summary_ws.cell(row=4, column=1).font = Font(bold=True)
            
            # Headers
            headers = ["Ticker", "Start Price", "Current Price", "Change", "Change %"]
            for col_idx, header in enumerate(headers, start=1):
                cell = summary_ws.cell(row=5, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
            
            # Data
            row_idx = 6
            for ticker, data in sorted(price_changes.items(), key=lambda x: x[1]['change_pct'], reverse=True):
                summary_ws.cell(row=row_idx, column=1, value=ticker)
                summary_ws.cell(row=row_idx, column=2, value=data["first_close"])
                summary_ws.cell(row=row_idx, column=2).number_format = '0.00'
                summary_ws.cell(row=row_idx, column=3, value=data["last_close"])
                summary_ws.cell(row=row_idx, column=3).number_format = '0.00'
                summary_ws.cell(row=row_idx, column=4, value=data["change"])
                summary_ws.cell(row=row_idx, column=4).number_format = '0.00'
                
                change_pct_cell = summary_ws.cell(row=row_idx, column=5, value=data["change_pct"]/100)
                change_pct_cell.number_format = '0.00%'
                if data["change_pct"] > 0:
                    change_pct_cell.font = Font(color="008800")
                elif data["change_pct"] < 0:
                    change_pct_cell.font = Font(color="CC0000")
                
                row_idx += 1
        
        book.save(file_path)
        logger.info("📊 Overview sheet updated with charts.")
    except Exception as e:
        logger.exception(f"Error updating overview sheet: {e}")

def main():
    try:
        # Initialize workbook
        ensure_workbook_exists(FILE_PATH, DEFAULT_TICKER)
        
        # Get list of tickers
        tickers = read_tickers(FILE_PATH)
        
        if not tickers:
            logger.warning("❌ No tickers found in the workbook.")
            return
        
        # Process each ticker
        for ticker in tickers:
            df = fetch_market_data(ticker)
            if not df.empty:
                append_data_to_sheet(df, ticker, FILE_PATH)
                update_ticker_status(ticker, FILE_PATH)
        
        # Update overview sheet
        update_overview_sheet(FILE_PATH, tickers, days=15)
        
        logger.info("Market data update completed successfully")
    except Exception as e:
        logger.exception(f"Unhandled error in main process: {e}")


if __name__ == "__main__":
    main()
    print("It worked!")